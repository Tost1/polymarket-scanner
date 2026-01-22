#!/usr/bin/env python3
"""
Polymarket Near-Certain Scanner
Tasks 1-8: Fetch markets + tags + exclude + threshold + flatten + 48h window + hours/URL + XLSX export
v1.0: Full pipeline, no testing limits
"""

import requests
import json
from datetime import datetime, timezone, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font


def fetch_all_markets(max_markets=None):
    """
    Fetch all markets from Polymarket Gamma API with pagination.
    Returns list of all market objects with tags included.
    
    Args:
        max_markets: Optional limit on total markets to fetch (for testing)
    """
    base_url = "https://gamma-api.polymarket.com/markets"
    all_markets = []
    offset = 0
    limit = 100  # Reasonable batch size
    
    print(f"Starting fetch from {base_url}")
    
    while True:
        params = {
            "limit": limit,
            "offset": offset,
            "closed": "false",
            "include_tag": "true"  # CRITICAL: includes tags in response
        }
        
        print(f"Fetching batch: offset={offset}, limit={limit}...", end=" ", flush=True)
        
        try:
            response = requests.get(base_url, params=params, timeout=30)
            print(f"Status: {response.status_code}", end=" ", flush=True)
            response.raise_for_status()
            markets = response.json()
            
            print(f"Got {len(markets)} markets")
            
            if not markets or len(markets) == 0:
                # No more markets to fetch
                print("No more markets, stopping pagination")
                break
            
            all_markets.extend(markets)
            
            # Stop if we've reached the max_markets limit
            if max_markets and len(all_markets) >= max_markets:
                print(f"Reached max_markets limit ({max_markets}), stopping")
                all_markets = all_markets[:max_markets]
                break
            
            # If we got fewer results than the limit, we've reached the end
            if len(markets) < limit:
                print(f"Received fewer than {limit} markets, stopping pagination")
                break
            
            offset += limit
            
        except requests.exceptions.RequestException as e:
            print(f"\nError fetching markets at offset {offset}: {e}")
            break
    
    return all_markets


def fetch_exclusion_tags():
    """
    Fetch specific exclusion tags (sports, esports, crypto) using individual lookups.
    Returns dict with found tags: {slug: {slug, label, id}}
    """
    base_url = "https://gamma-api.polymarket.com/tags/slug"
    exclusion_slugs = ['sports', 'esports', 'crypto']
    found_tags = {}
    
    print(f"Fetching exclusion tags...")
    
    for slug in exclusion_slugs:
        url = f"{base_url}/{slug}"
        print(f"  Checking '{slug}'...", end=" ", flush=True)
        
        try:
            response = requests.get(url, timeout=30)
            
            if response.status_code == 200:
                tag_data = response.json()
                found_tags[slug] = {
                    'slug': tag_data.get('slug', slug),
                    'label': tag_data.get('label', ''),
                    'id': tag_data.get('id', '')
                }
                print(f"✓ Found (label: '{tag_data.get('label', 'N/A')}', ID: {tag_data.get('id')})")
            elif response.status_code == 404:
                print(f"✗ Not found (404)")
            else:
                print(f"✗ Error (status {response.status_code})")
                
        except requests.exceptions.RequestException as e:
            print(f"✗ Error: {e}")
    
    return found_tags


def exclude_by_tags(markets, exclusion_tags):
    """
    Exclude markets that have any of the exclusion tags.
    Tags are now included in market objects via include_tag=true.
    Returns (filtered_markets, excluded_markets).
    """
    # Build set of exclusion tag IDs and slugs/labels (case-insensitive)
    exclusion_ids = set(int(tag['id']) for tag in exclusion_tags.values() if tag['id'])
    exclusion_slugs = set(tag['slug'].lower() for tag in exclusion_tags.values())
    exclusion_labels = set(tag['label'].lower() for tag in exclusion_tags.values())
    
    filtered = []
    excluded = []
    
    for market in markets:
        # Get tags array from market (now included with include_tag=true)
        market_tags = market.get('tags', [])
        
        # Check if any market tag matches exclusion criteria
        has_excluded_tag = False
        matched_tags = []
        
        for tag in market_tags:
            tag_id = tag.get('id')
            tag_slug = tag.get('slug', '').lower()
            tag_label = tag.get('label', '').lower()
            
            # Match on ID, slug, or label
            if (tag_id in exclusion_ids or 
                tag_slug in exclusion_slugs or 
                tag_label in exclusion_labels):
                has_excluded_tag = True
                matched_tags.append(tag)
        
        if has_excluded_tag:
            market['_matched_tags'] = matched_tags  # Store for debugging
            excluded.append(market)
        else:
            filtered.append(market)
    
    return filtered, excluded


def exclude_by_keywords(markets):
    """
    Exclude markets matching esports keywords (backup filter).
    Searches in: question, event.title, category, subcategory (case-insensitive).
    Returns (filtered_markets, excluded_markets).
    """
    esports_keywords = [
        'esports',
        'cs2',
        'cs:go',
        'dota',
        'league of legends',
        'valorant',
        'overwatch'
    ]
    
    filtered = []
    excluded = []
    
    for market in markets:
        # Extract searchable fields
        question = market.get('question', '').lower()
        event_title = market.get('event', {}).get('title', '').lower() if isinstance(market.get('event'), dict) else ''
        category = market.get('category', '').lower()
        subcategory = market.get('subcategory', '').lower()
        
        # Combine all searchable text
        searchable_text = f"{question} {event_title} {category} {subcategory}"
        
        # Check if any keyword matches
        matched_keywords = []
        for keyword in esports_keywords:
            if keyword in searchable_text:
                matched_keywords.append(keyword)
        
        if matched_keywords:
            market['_matched_keywords'] = matched_keywords  # Store for debugging
            excluded.append(market)
        else:
            filtered.append(market)
    
    return filtered, excluded


def apply_price_threshold(markets, threshold=0.95):
    """
    Keep markets where ANY outcome has price >= threshold OR <= (1 - threshold).
    
    For binary markets (["Yes", "No"]):
    - Near-certain YES: yes_price >= 0.95
    - Near-certain NO: yes_price <= 0.05 (equivalent to no_price >= 0.95)
    
    For multi-outcome markets:
    - Keep if ANY outcome price >= 0.95
    
    Market data structure:
    - outcomes: stringified JSON array like '["Yes", "No"]' or '["Outcome A", "Outcome B", "Outcome C"]'
    - outcomePrices: stringified JSON array like '["0.65", "0.35"]'
    
    Returns (markets_meeting_threshold, markets_below_threshold).
    """
    meeting_threshold = []
    below_threshold = []
    
    for market in markets:
        # Parse outcomes and prices (they're stringified JSON)
        outcomes_str = market.get('outcomes', '[]')
        prices_str = market.get('outcomePrices', '[]')
        
        try:
            outcomes = json.loads(outcomes_str)
            prices = json.loads(prices_str)
            
            # Convert price strings to floats
            prices = [float(p) for p in prices]
            
            # Determine if market is binary
            is_binary = outcomes == ["Yes", "No"]
            
            if is_binary:
                # Binary market: check YES price (prices[0])
                yes_price = prices[0]
                
                # Near-certain YES: yes_price >= 0.95
                # Near-certain NO: yes_price <= 0.05
                if yes_price >= threshold or yes_price <= (1 - threshold):
                    market['_is_binary'] = True
                    market['_yes_price'] = yes_price
                    market['_no_price'] = prices[1] if len(prices) > 1 else (1 - yes_price)
                    meeting_threshold.append(market)
                else:
                    below_threshold.append(market)
            else:
                # Multi-outcome market: check if ANY outcome >= threshold
                max_price = max(prices) if prices else 0.0
                
                if max_price >= threshold:
                    market['_is_binary'] = False
                    market['_outcomes'] = outcomes
                    market['_prices'] = prices
                    meeting_threshold.append(market)
                else:
                    below_threshold.append(market)
                    
        except (json.JSONDecodeError, ValueError, IndexError) as e:
            # Skip markets with malformed data
            print(f"Warning: Skipping market due to parse error: {e}")
            below_threshold.append(market)
    
    return meeting_threshold, below_threshold


def flatten_multi_outcome_markets(markets, threshold=0.95):
    """
    Flatten markets to 1 row per outcome.
    
    For binary markets:
    - 1 row with outcome='YES' or 'NO' depending on which side is >= 0.95
    
    For multi-outcome markets:
    - 1 row per outcome where price >= threshold
    
    Each row contains:
    - market: original market object
    - outcome: outcome name
    - yes_price: outcome price
    - no_price: None (for multi-outcome) or complementary price (for binary)
    - certainty_side: 'YES', 'NO', or outcome name
    - is_binary: boolean flag
    """
    rows = []
    
    for market in markets:
        is_binary = market.get('_is_binary', False)
        
        if is_binary:
            # Binary market: create 1 row for the near-certain side
            yes_price = market['_yes_price']
            no_price = market['_no_price']
            
            if yes_price >= threshold:
                # Near-certain YES
                rows.append({
                    'market': market,
                    'outcome': 'YES',
                    'yes_price': yes_price,
                    'no_price': no_price,
                    'certainty_side': 'YES',
                    'is_binary': True
                })
            elif yes_price <= (1 - threshold):
                # Near-certain NO (yes_price <= 0.05 means no_price >= 0.95)
                rows.append({
                    'market': market,
                    'outcome': 'NO',
                    'yes_price': yes_price,
                    'no_price': no_price,
                    'certainty_side': 'NO',
                    'is_binary': True
                })
        else:
            # Multi-outcome market: create 1 row per outcome where price >= threshold
            outcomes = market.get('_outcomes', [])
            prices = market.get('_prices', [])
            
            for outcome, price in zip(outcomes, prices):
                if price >= threshold:
                    rows.append({
                        'market': market,
                        'outcome': outcome,
                        'yes_price': price,
                        'no_price': None,  # No complement for multi-outcome
                        'certainty_side': outcome,
                        'is_binary': False
                    })
    
    return rows


def apply_time_window(rows, window_hours=48):
    """
    Filter rows to include only markets with Resolve_DateTime in [now, now + window_hours].
    Also calculate Hours_Remaining and construct Market_URL for each row.
    
    Returns:
        - in_window: rows within the time window
        - outside_window: rows outside the time window
        - now: current datetime
        - window_end: end of time window
        - min_date: earliest endDate seen
        - max_date: latest endDate seen
    """
    now = datetime.now(timezone.utc)
    window_end = now + timedelta(hours=window_hours)
    
    in_window = []
    outside_window = []
    all_dates = []
    
    for row in rows:
        market = row['market']
        
        # Parse endDate (ISO format: "2025-01-23T14:00:00Z")
        end_date_str = market.get('endDate', '')
        
        try:
            # Parse ISO datetime string
            end_date = datetime.fromisoformat(end_date_str.replace('Z', '+00:00'))
            all_dates.append(end_date)
            
            # Check if within window
            if now <= end_date <= window_end:
                # Calculate hours remaining
                time_remaining = end_date - now
                hours_remaining = time_remaining.total_seconds() / 3600
                
                # Construct market URL using market slug
                # URL pattern: https://polymarket.com/market/{market_slug}
                market_slug = market.get('slug', '')
                market_url = f"https://polymarket.com/market/{market_slug}" if market_slug else ""
                
                # Add calculated fields to row
                row['resolve_datetime'] = end_date
                row['hours_remaining'] = hours_remaining
                row['market_url'] = market_url
                
                in_window.append(row)
            else:
                outside_window.append(row)
                
        except (ValueError, AttributeError) as e:
            # Skip markets with malformed endDate
            print(f"Warning: Skipping market due to date parse error: {e}")
            outside_window.append(row)
    
    # Calculate min/max dates seen
    min_date = min(all_dates) if all_dates else None
    max_date = max(all_dates) if all_dates else None
    
    return in_window, outside_window, now, window_end, min_date, max_date


def export_to_xlsx(rows, output_path='markets_raw.xlsx'):
    """
    Export rows to XLSX file sorted by Resolve_DateTime.
    
    Column order (from spec.md):
    1. Event_Title
    2. Market_Question
    3. Outcome
    4. YES_Price
    5. NO_Price
    6. Certainty_Side (YES / NO / OUTCOME)
    7. Category
    8. Subcategory
    9. Volume
    10. Liquidity
    11. Resolve_DateTime
    12. Hours_Remaining
    13. Market_URL (clickable)
    14. AI_Confidence (empty)
    15. AI_Rationale (empty)
    """
    # Sort rows by resolve_datetime (earliest first)
    sorted_rows = sorted(rows, key=lambda r: r['resolve_datetime'])
    
    # Create workbook and active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Markets"
    
    # Define headers (exact order from spec)
    headers = [
        'Event_Title',
        'Market_Question',
        'Outcome',
        'YES_Price',
        'NO_Price',
        'Certainty_Side',
        'Category',
        'Subcategory',
        'Volume',
        'Liquidity',
        'Resolve_DateTime',
        'Hours_Remaining',
        'Market_URL',
        'AI_Confidence',
        'AI_Rationale'
    ]
    
    # Write headers
    ws.append(headers)
    
    # Make headers bold
    for cell in ws[1]:
        cell.font = Font(bold=True)
    
    # Write data rows
    for row in sorted_rows:
        market = row['market']
        
        # Extract event title (handle both dict and string)
        event = market.get('event')
        if isinstance(event, dict):
            event_title = event.get('title', '')
        else:
            event_title = str(event) if event else ''
        
        # Format NO_Price (empty for multi-outcome markets)
        no_price_display = row['no_price'] if row['no_price'] is not None else ''
        
        # Format resolve datetime
        resolve_datetime_str = row['resolve_datetime'].strftime('%Y-%m-%d %H:%M:%S UTC')
        
        # Build row data (exact column order)
        row_data = [
            event_title,
            market.get('question', ''),
            row['outcome'],
            row['yes_price'],
            no_price_display,
            row['certainty_side'],
            market.get('category', ''),
            market.get('subcategory', ''),
            market.get('volume', ''),
            market.get('liquidity', ''),
            resolve_datetime_str,
            round(row['hours_remaining'], 2),
            row['market_url'],
            '',  # AI_Confidence (empty)
            ''   # AI_Rationale (empty)
        ]
        
        ws.append(row_data)
    
    # Set column widths for better readability
    column_widths = {
        'A': 30,  # Event_Title
        'B': 60,  # Market_Question
        'C': 30,  # Outcome
        'D': 12,  # YES_Price
        'E': 12,  # NO_Price
        'F': 15,  # Certainty_Side
        'G': 20,  # Category
        'H': 20,  # Subcategory
        'I': 15,  # Volume
        'J': 15,  # Liquidity
        'K': 22,  # Resolve_DateTime
        'L': 15,  # Hours_Remaining
        'M': 80,  # Market_URL
        'N': 15,  # AI_Confidence
        'O': 40   # AI_Rationale
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Make URLs clickable using HYPERLINK formula (OnlyOffice-safe)
    # Column M = 13 = Market_URL
    url_col_idx = 13
    for row_idx in range(2, len(sorted_rows) + 2):  # Start from row 2 (skip header)
        cell = ws.cell(row=row_idx, column=url_col_idx)
        url = cell.value
        if url:
            # Use HYPERLINK formula: =HYPERLINK("url","open")
            cell.value = f'=HYPERLINK("{url}","open")'
            cell.font = Font(color="0000FF", underline="single")
    
    # Save workbook
    wb.save(output_path)
    print(f"Exported {len(sorted_rows)} rows to {output_path}")


def main():
    """v1.0 Full Pipeline: No testing limits"""
    print("="*60)
    print("POLYMARKET SCANNER v1.0 - FULL PIPELINE")
    print("="*60)
    print()
    
    # Fetch exclusion tags
    exclusion_tags = fetch_exclusion_tags()
    print()
    
    # Fetch ALL markets (no limit)
    markets = fetch_all_markets()
    print(f"Total markets fetched: {len(markets)}")
    print()
    
    # Apply tag-based exclusions
    print("Applying tag-based exclusions...")
    after_tags, excluded_by_tags = exclude_by_tags(markets, exclusion_tags)
    print(f"After tag exclusions: {len(after_tags)} markets remaining ({len(excluded_by_tags)} excluded)")
    print()
    
    # Apply keyword-based exclusions
    print("Applying keyword-based exclusions...")
    after_keywords, excluded_by_keywords = exclude_by_keywords(after_tags)
    print(f"After keyword exclusions: {len(after_keywords)} markets remaining ({len(excluded_by_keywords)} excluded)")
    print()
    
    # Apply price threshold
    print("Applying 0.95 price threshold...")
    final_markets, below_threshold = apply_price_threshold(after_keywords, threshold=0.95)
    print(f"Markets meeting threshold: {len(final_markets)} ({len(below_threshold)} below threshold)")
    print()
    
    # Flatten markets
    print("Flattening markets...")
    flattened_rows = flatten_multi_outcome_markets(final_markets, threshold=0.95)
    print(f"Total flattened rows: {len(flattened_rows)}")
    
    # Count binary YES/NO rows
    binary_yes = sum(1 for r in flattened_rows if r.get('is_binary') and r.get('certainty_side') == 'YES')
    binary_no = sum(1 for r in flattened_rows if r.get('is_binary') and r.get('certainty_side') == 'NO')
    multi_outcome = sum(1 for r in flattened_rows if not r.get('is_binary'))
    print(f"  - Binary YES: {binary_yes}")
    print(f"  - Binary NO: {binary_no}")
    print(f"  - Multi-outcome: {multi_outcome}")
    print()
    
    # Apply 48-hour time window
    print("Applying 48-hour time window...")
    rows_in_window, rows_outside, now, window_end, min_date, max_date = apply_time_window(flattened_rows, window_hours=48)
    print(f"Rows in 48h window: {len(rows_in_window)} ({len(rows_outside)} outside window)")
    print()
    
    # Export to XLSX
    if rows_in_window:
        print("Exporting to markets_raw.xlsx...")
        export_to_xlsx(rows_in_window, output_path='markets_raw.xlsx')
        print()
        print("="*60)
        print("EXPORT COMPLETE")
        print("="*60)
        print(f"File: markets_raw.xlsx")
        print(f"Rows exported: {len(rows_in_window)}")
        print()
        print("Next steps:")
        print("  1. Open markets_raw.xlsx in OnlyOffice")
        print("  2. Verify column order matches spec.md")
        print("  3. Verify sort by Resolve_DateTime (earliest first)")
        print("  4. Test clickable URLs")
    else:
        print("="*60)
        print("NO MARKETS FOUND")
        print("="*60)
        print("No markets found within 48-hour window. No file exported.")


if __name__ == "__main__":
    main()